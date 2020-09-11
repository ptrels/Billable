import os, openpyxl, datetime, time, pprint, json, csv
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from tkinter import *
from tkinter import ttk, Text, Tk, Menu, scrolledtext, messagebox, filedialog
from datetime import date, timedelta

# creates tkinter window
window = Tk()

window.title("Timer - Billable Hours / Tasks")
window.geometry("500x400")

class Stop_Watch(Frame):  
    # creates a stopwatch frame widget                                                              
    def __init__(self, parent = None, **kwargs):        
        Frame.__init__(self, parent, kwargs)
        self.start = 0.0
        self.start_time = datetime.datetime.now()
        self.date = self.start_time.strftime("%m/%d/%Y")
        self.stop_time = datetime.datetime.now()
        self.end_time = datetime.datetime.now()
        self.all_vars = []
        self.elapsed_time = 0.0
        self.running = False
        self.time_str = StringVar()
        self.time_hours = StringVar()
        self.make_widgets()      

    def make_widgets(self):                         
        # make the time label
        l = Label(self, textvariable = self.time_str, font=("bold", 14))
        self.set_time(self.elapsed_time)
        l.pack(fill = X, expand = NO, pady = 2, padx = 2)                      
    
    def update(self): 
        # update the label with elapsed time
        self.elapsed_time = time.time() - self.start
        self.set_time(self.elapsed_time)
        self.timer = self.after(50, self.update)
    
    def set_time(self, elap):
        # set the time string to Minutes:Seconds:Hundreths
        hours_1 = int(elap / 3600.0)
        minutes_1 = int((elap - hours_1*3600.0) / 60)
        seconds_1 = int(elap - hours_1*3600.0 - minutes_1*60.0)
        hseconds_1 = int((elap - hours_1*3600.0 - minutes_1*60.0 - seconds_1)*100)
        minutes = int(elap/60)
        seconds = int(elap - minutes*60.0)
        hseconds = int((elap - minutes*60.0 - seconds)*100) 
        self.time_str.set("%02d:%02d:%02d" % (minutes, seconds, hseconds))
        self.time_hours.set("%02d:%02d:%02d:%02d" % (hours_1, minutes_1, seconds_1, hseconds_1))
        
    def Start(self):                                                     
        # start the stopwatch, ignore if running
        if not self.running:            
            self.start = time.time() - self.elapsed_time
            self.start_time = datetime.datetime.now()
            self.update()
            self.running = True        
    
    def Stop(self):                                    
        # stop the stopwatch, ignore if stopped
        if self.running:
            self.after_cancel(self.timer) # stops tkinter after running function (pause clock)
            self.stop_time = datetime.datetime.now()
            self.elapsed_time = time.time() - self.start    
            self.set_time(self.elapsed_time)
            self.Store()
            self.running = False
    
    def Reset(self):                                  
        # reset the stopwatch
        self.start = time.time()         
        self.elapsed_time = 0.0    
        self.set_time(self.elapsed_time)
        self.start_time = datetime.datetime.now()
        self.stop_time = datetime.datetime.now()

    def Store(self):
        # compiles time variables from stopwatch, then returns all as variable to be saved
        self.date = self.start_time.strftime("%m/%d/%Y")
        self.client = combo_client.get()
        self.task = combo_task.get()
        self.notes = notes.get(1.0, END)
        self.start_time = self.start_time.strftime("%I:%M:%S %p")
        self.stop_time = self.stop_time.strftime("%I:%M:%S %p")
        ending_time = self.time_hours.get()
        self.all_vars = [self.date, self.client, self.task, self.notes, self.start_time, self.stop_time, self.elapsed_time, ending_time]
        return self.all_vars

direct_called_yes = False
data_stored_yes = False

# see if data_stored, sets global data_stored_yes variable
def data_stored(value):
    global data_stored_yes
    data_stored_yes = value

# see if directory set, sets global direct_called_yes variable
def direct_called(value):
    global direct_called_yes
    direct_called_yes = value
        

class Checks:
    def __init__(self, folder_path, file_name):
        self.file_path = "" # where .py file saved
        self.folder_path = folder_path # Directory folder
        self.file_name = file_name # file name (Directory.xlsx or Directory.txt)
        self.client_names = []
        self.tasks_types = []
        self.lists = []

    # sets directory path to where .py file saved     
    def set_direct(self):
        self.file_path = os.getcwd()
        self.folder_path = os.path.join(self.file_path, self.folder_path)

    # checks directory path, or creates, and changes path to Directory folder
    def check_direct(self):
        if os.path.exists(self.folder_path):
            os.chdir(self.folder_path)
        else:
            os.makedirs("Directory")
            os.chdir(self.folder_path)

    # function to check if excel file exists
    # if not, then creates
    def check_excel(self):
        if os.path.exists(os.path.join(self.folder_path, self.file_name + ".xlsx")):
            pass
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Directory"
            headers = ["Date", "Client", "Task", "Notes", "Start Time", "End Time", "Time Elapsed (Mins: Sec: MS)", "Total Time (Hrs : Mins : Sec : MS)"]
            ws.append(headers)
            bold_12_font = Font(size = 12, bold = True)
            for cell in ws["1:1"]:
                cell.font = bold_12_font
            for cell in ws["A:A"]:
                cell.number_format = "M/D/YYYY"
            column = 1
            max_cols = ws.max_column + 1
            while column < max_cols:
                i = get_column_letter(column)
                ws.column_dimensions[i].width = 15
                column = column + 1
            wb.save(filename = self.file_name + ".xlsx")
            wb.close()

    # checks if csv file exists
    # if not, then creates
    def check_csv(self):
        if os.path.exists(os.path.join(self.folder_path, self.file_name + ".csv")):
            pass
        else:
            headers = ["Date", "Client", "Task", "Notes", "Start Time", "End Time", "Total Time (Hrs : Mins : Sec : MS)"]
            with open("Directory.csv", "w+", newline = "") as out_csv:
                writer = csv.writer(out_csv)
                writer.writerow(headers)
            
    # Checks if Directory.json exists
    # if it does, then loads list from Directory.json
    def ct_list(self):
        path = os.path.join("Directory.json")
        if os.path.exists(path):
            try:
                fp_load = open(path, "r")
                lists = json.loads(fp_load.read())
                client_names = lists[0]
                tasks_types = lists[1]
                self.client_names = client_names
                self.tasks_types = tasks_types
                fp_load.close()
            except IndexError:
                client_names = ["Client_1, Client_2, Client_3"]
                tasks_types = ["Task_1, Task_2, Task_3"]
                lists = [client_names, tasks_types]
                fp_up = open("Directory.json", "w+")
                fp_up.write(json.dumps(lists, indent = 4))
                self.client_names = client_names
                self.tasks_types = tasks_types
                fp_up.close()

        # if Directory.json does not exists, then creates it with dummy variables in lists
        else:
            client_names = ["Client_1", "Client_2", "Client_3"]
            self.client_names = client_names
            tasks_types = ["Research", "Writing", "Memo"]
            self.tasks_types = tasks_types
            lists = [client_names, tasks_types]
            self.lists = lists
            fp = open("Directory.json", "w+")
            fp.write(json.dumps(lists, indent = 4))
            fp.close()

    # appends lists if user types in different name on main window combobox
    # checks to make sure not already on the list before appends it
    def lists_append(self):
        client_names = self.client_names
        tasks_types = self.tasks_types
        lists = self.lists
        path = os.path.join("Directory.json")
        
        if combo_client.get() not in client_names:
            client_names.append(combo_client.get())
            self.client_names = client_names
            if combo_task.get() not in tasks_types:
                tasks_types.append(combo_task.get())
                self.tasks_types = tasks_types
                lists = [client_names, tasks_types]
                fp = open(path, "w+")
                fp.write(json.dumps(lists, indent = 4))
                fp.close()
            else:
                lists = [client_names, tasks_types]
                fp = open(path, "w+")
                fp.write(json.dumps(lists, indent = 4))
                fp.close()
        elif combo_client.get() in client_names:
            if combo_task.get() not in tasks_types:
                tasks_types.append(combo_task.get())
                self.tasks_types = tasks_types
                lists = [client_names, tasks_types]
                fp = open(path, "w+")
                fp.write(json.dumps(lists, indent = 4))
                fp.close()
            else:
                pass       
        else:
            pass
    
    # creates edit list toplevel window and functions to edit lists in program
    def edit_lists(self, list_value):
        client_names = self.client_names
        tasks_types = self.tasks_types
        lists = self.lists
        path = os.path.join("Directory.json")        

        # creates Toplevel window
        editor_win = Toplevel(window)
        editor_win.title("Edit List")

        # Listbox
        listbox = Listbox(editor_win, selectmode = EXTENDED)
        listbox.grid(padx = 50, pady = 5)

        # if list_value (argument) is client_names, then opens that list
        if list_value is client_names:
            lbl_5 = Label(editor_win, text="Client Names", font=("Arial Bold", 10))
            lbl_5.grid(padx = 5, pady = 5)
            for item in client_names:
                listbox.insert(END, item)
        # if list_value (argument) is tasks_types, then opens that list
        elif list_value is tasks_types:
            lbl_5 = Label(editor_win, text="Task Types", font=("Arial Bold", 10))
            lbl_5.grid(padx = 5, pady = 5)
            for item in tasks_types:
                listbox.insert(END, item)
        
        # deletes selection
        def delete_list(self):
            selection = self.curselection()
            for index in reversed(selection):
                self.delete(index)

        # adds text entry to list
        def add_list(self, list_value):
            if insert_text.get() in list_value:
                messagebox.showerror("Error", "This entry is already on the list", parent = editor_win)
                pass
            else:
                self.insert(END, insert_text.get())
                self.see(END)
                insert_text.delete(0, END)

        # function to save list
        def save_list(win, list_value):
            client_names = self.client_names
            tasks_types = self.tasks_types
            final_list = list(win.get(0, END))
            if list_value is client_names:
                client_names = final_list
                self.client_names = client_names
            elif list_value is tasks_types:
                tasks_types = final_list
                self.tasks_types = tasks_types
            lists = [client_names, tasks_types]
            fp = open("Directory.json", "w+")
            fp.write(json.dumps(lists, indent = 4))
            fp.close()
            list_refresh()

        # calls save function and exits when close top window
        def exit_list():
            save_list(listbox, list_value)
            editor_win.destroy()
            combo_client.configure(values = self.client_names)
            combo_task.configure(values = self.tasks_types)

        insert_text = Entry(editor_win, width = 20)
        insert_text.grid(padx = 30, pady = 5)
        btn_7 = Button(editor_win, text = "Add", fg = "blue", bg = "white", command = lambda: add_list(listbox, list_value))
        btn_7.grid(padx = 30, pady = 5)
        btn_6 = Button(editor_win, text = "Delete", fg = "red", bg = "white", command = lambda: delete_list(listbox))
        btn_6.grid(padx = 30, pady = 5)
        
        # bind enter key press to add_list
        editor_win.bind("<Return>", lambda event = None: btn_7.invoke())
        # bind delete key press to delete_list
        editor_win.bind("<Delete>", lambda event = None: btn_6.invoke())
        
        editor_win.protocol("WM_DELETE_WINDOW", exit_list)       
        

check = Checks("Directory", "Directory")

# refreshes client_names and tasks_types list to whatever stored
def list_refresh():
    global client_names
    global tasks_types
    client_names = check.client_names
    tasks_types = check.tasks_types

# checks directory, folder, and files
def checker():
    check.set_direct()
    check.check_direct()
    check.check_excel()
    check.check_csv()
    check.ct_list()
    direct_called(True)
    messagebox.showwarning("Directory Path Set", check.folder_path)

# clear text entry widgets function
def clear_text():
    if not data_stored_yes:
        yes_no_4 = messagebox.askyesno("Data Warning", "Data has not been saved. Do you want to save the entry before clearing the text?")
        if yes_no_4:
            store_data(store = True)
        else:
            data_stored(False)
            sw.Reset()
            combo_client.set("")
            combo_task.set("")
            notes.delete(1.0, END)
            combo_client.focus_force()
    else:
        data_stored(False)
        sw.Reset()
        combo_client.set("")
        combo_task.set("")
        notes.delete(1.0, END)
        combo_client.focus_force()
    
# store data in excel function
def store_data(store = False):
    if data_stored_yes:
        yes_no_5 = messagebox.askyesno("Warning", "This entry may have already been logged. Save a new entry anyways?")
        if yes_no_5:
            data_stored(False)
            store_data()
    else:
        if direct_called_yes:
            try:
                # appends excel file and saves as new entry
                check.lists_append()
                wb = openpyxl.load_workbook("Directory.xlsx")
                ws = wb.active
                bold_12_font = Font(size = 12, bold = True)
                ws.append(sw.all_vars)
                wb.save("Directory.xlsx")

                # appends csv file and saves as new entry
                with open("Directory.csv", "a+", newline = "") as append_csv:
                    append = csv.writer(append_csv)
                    append.writerow(sw.all_vars)
                data_stored(True)

                if store == True:
                    # This is for new entry button storing text before clearing text
                    messagebox.showwarning("Data Stored", "Data saved successfully.")
                    clear_text()
                elif store == "exit":
                    # this is for saving before exiting the program
                    messagebox.showwarning("Data Stored", "Data saved successfully.")
                    window.destroy()
                else:
                    # this is for save button storing text
                    yes_no_3 = messagebox.askyesno("Data Stored", "Data saved successfully. Do you want to clear the text?")
                    if yes_no_3:
                        combo_client.configure(values = check.client_names)
                        combo_task.configure(values = check.tasks_types)
                        clear_text()
                    else:
                        data_stored(True)
                        pass

            except PermissionError:
                messagebox.showerror("Error", "Please close the Directory Excel or CSV file before continuing.")
                store_data()
            except FileNotFoundError:
                checker()
                store_data()
            except NameError:
                pass       

        else:
            checker()
            store_data()
    
# new entry function    
def new_entry():
    yes_no = messagebox.askyesnocancel("New Entry", "Do you want to start a new entry")
    if yes_no is True and data_stored_yes is False:
        store_data()
        clear_text()
    else:
        pass
    
# exit command function
def exit_prog():
    try:
        if sw.elapsed_time == 0.0:
            window.destroy()
        elif data_stored_yes:
            window.destroy()
        else:
            yes_no_1 = messagebox.askyesnocancel("Exit", "Do you want to save this entry and exit the program?\nNo will exit without saving.")
            if yes_no_1 is True and data_stored_yes is False:
                store_data(store = "exit")
            elif yes_no_1 is None:
                pass
            else:
                window.destroy()
    except:
        print("Error")
        window.destroy()

# about me messagebox - menu
def about():
    about_text ="""
Billable

Created by: Peter Trelenberg

11/6/2019
Version: 1.0
email: pt3791a@american.edu\n
""" + "File Path:\n" + check.file_path + "\n\nFolder Path:\n" + check.folder_path
    messagebox.showinfo("About Billable", about_text)

# help messagebox - menu
def help_info():
    help_win = Toplevel(window)
    help_win.title("Help")
    
    help_text = """Program to track and store the amount of time spent working for a client / on a project.

Billable creates a "Directory" folder and saves entries in the "Directory.xlsx" file.

Both the Client list and Task list are stored within the "Directory.py" file saved in the "Directory" folder.

Billable has three main sections within the Main Window:

Top Section - Timer Function
Left Section - Variable Inputs
Right Section - Actionable Buttons

Editable variable inputs include: Client, Task, and Notes.

The Client or Task list can be changed in two ways:

1) click the associated button in the right frame, add / delete item and close the window to save.
2) type in a new item and it will be saved along with the entry.

Actionable Buttons include:

Save: Saves current entry and appends the entry to the "Directory.xlsx" file.
New Entry: Asks user if they want to save the current entry and then clears the text for a new entry.
Edit Client List: Opens new window to allow user to edit Client list.
Edit Task List: Opens new window to allow user to edit Task list.
Exit: Asks user if they want to save and then exits the program.
"""

    help_lbl = Label(help_win, text="Billable", font=("Arial Bold", 12))
    help_lbl.grid()
    help_txt_1 = Text(help_win, wrap = "word")
    help_txt_1.grid(padx = 10, pady = 10)
    help_txt_1.insert(1.0, help_text)

    # tags to format text
    help_txt_1.tag_add("highlight", "3.19", "3.37")
    help_txt_1.tag_add("highlight", "3.63", "3.79")
    help_txt_1.tag_add("highlight", "5.57", "5.71")
    help_txt_1.tag_add("highlight", "5.90", "5.108")
    help_txt_1.tag_add("highlightline", "9.0", "12.0")
    help_txt_1.tag_add("highlightline", "16.0", "19.0")
    help_txt_1.tag_add("highlightline", "22.0", "27.0")
    help_txt_1.tag_add("bold", "22.0", "22.5")
    help_txt_1.tag_add("bold", "23.0", "23.10")
    help_txt_1.tag_add("bold", "24.0", "24.17")
    help_txt_1.tag_add("bold", "25.0", "25.15")
    help_txt_1.tag_add("bold", "26.0", "26.5")
    help_txt_1.tag_configure("highlight", foreground = "red")
    help_txt_1.tag_configure("highlightline", lmargin1 = 20, lmargin2 = 20, foreground = "blue")
    help_txt_1.tag_configure("bold", underline = 1, foreground = "green")
    help_txt_1.config(state = DISABLED)

checker()

# ask to save before closing using x
window.protocol("WM_DELETE_WINDOW", exit_prog)

# Menu items
menu = Menu(window)
new_item = Menu(menu, tearoff = 0)
help_menu = Menu(menu, tearoff = 0)

menu.add_cascade(label = "File", menu = new_item)
menu.add_cascade(label = "Help Menu", menu = help_menu)

help_menu.add_command(label = "About Billable", command = about)
help_menu.add_command(label = "Help", command = help_info)
new_item.add_command(label = "Directory", command = checker)
new_item.add_command(label = "New Entry", command = clear_text)
new_item.add_command(label = "Exit", command = exit_prog)
window.config(menu = menu)

# create all of the main containers
left_frame = Frame(window, width = 300, height = 300)
right_frame = Frame(window, width = 200, height = 300)
top_frame = Frame(window, width = 500, height = 100, pady = 10)

# layout all of the main containers
left_frame.grid_propagate(False)
right_frame.grid_propagate(False)
top_frame.grid_propagate(False)

left_frame.grid(column = 0, row = 1)
right_frame.grid(column = 1, row = 1)
top_frame.grid(column = 0, row = 0, columnspan = 2)

## create the widgets for the left frame
lbl = Label(left_frame, text="Fill out the information and click start to begin timer", font=("Arial Bold", 8))

# client combobox
lbl_1 = Label(left_frame, text="Client:", font=("Arial Bold", 8))
combo_client = ttk.Combobox(left_frame, values = check.client_names)
combo_client.current(0) #set the selected item

# task entry box
lbl_2 = Label(left_frame, text="Task:", font=("Arial Bold", 8))
combo_task = ttk.Combobox(left_frame, values = check.tasks_types)
combo_task.current(0) #set the selected item

# notes entry box
lbl_3 = Label(left_frame, text="Notes:", font=("Arial Bold", 8))
notes = scrolledtext.ScrolledText(left_frame, width = 30, height = 6)

## layout widgets for the left frame
lbl.grid(column = 0, row = 0, columnspan = 2, pady = 5)
lbl_1.grid(column = 0, row = 1, pady = 5)
combo_client.grid(column = 1, row = 1, pady = 5)
lbl_2.grid(column = 0, row = 2, pady = 5)
combo_task.grid(column = 1, row = 2, pady = 5)
lbl_3.grid(column = 0, row = 3, pady = 5)
notes.grid(column = 1, row = 3, pady = 5)

## create the widgets for the right frame
list_refresh()

# actionable buttons
btn_5 = Button(right_frame, text = "Save", fg = "green4", bg = "white", command = store_data)
btn_1 = Button(right_frame, text = "New Entry", fg = "RoyalBlue2", bg = "white", command = clear_text)
btn_3 = Button(right_frame, text = "Edit Client List", fg = "dark orange", bg = "white", command = lambda: check.edit_lists(client_names))
btn_10 = Button(right_frame, text = "Edit Task List", fg = "gold", bg = "white", command = lambda: check.edit_lists(tasks_types))
btn_4 = Button(right_frame, text = "Exit", fg = "red3", bg = "white", command = exit_prog)


## layout widget for the right frame
btn_5.grid(columnspan = 3, padx = 30, pady = (30, 3))
btn_1.grid(columnspan = 3, padx = 30, pady = 3)
btn_3.grid(columnspan = 3, padx = 30, pady = 3)
btn_10.grid(columnspan = 3, padx = 30, pady = 3)
btn_4.grid(columnspan = 3, padx = 30, pady = 3)


## create and layout widgets for top frame
sw = Stop_Watch(top_frame)
sw.pack(side = TOP)

start_btn = Button(top_frame, text="Start", fg = "green", font = ("bold", 12), command = sw.Start)
start_btn.pack(side = LEFT)

stop_btn = Button(top_frame, text="Stop", fg = "red", font = ("bold", 12), command = sw.Stop)
stop_btn.pack(side = LEFT)

reset_btn = Button(top_frame, text="Reset", fg = "blue", font = ("bold", 12), command = sw.Reset)
reset_btn.pack(side = LEFT)

combo_client.focus_force()

def minimize():
    if sw.running:
        window.iconify()
    
        top_wind = Toplevel(window)
        top_wind.title("Timer")
                    
        # make the time label
        lbl_clock = Label(top_wind, textvariable = sw.time_str, font=("bold", 14))
        lbl_clock.pack(fill = X, expand = NO, pady = 2, padx = 2)
        
        start_btn_1 = Button(top_wind, text="Start", fg = "green", font = ("bold", 12), command = sw.Start)
        start_btn_1.pack(side = LEFT)

        stop_btn_1 = Button(top_wind, text="Stop", fg = "red", font = ("bold", 12), command = sw.Stop)
        stop_btn_1.pack(side = LEFT)

        reset_btn_1 = Button(top_wind, text="Reset", fg = "blue", font = ("bold", 12), command = sw.Reset)
        reset_btn_1.pack(side = LEFT)

        window.unbind("<Unmap>")
        top_wind.focus_force()

        def close():
            top_wind.unbind("<Unmap>")
            top_wind.destroy()
            window.deiconify()
            window.focus_force()
            window.bind("<Unmap>", lambda event = None: minimize())
            pass

        top_wind.bind("<Unmap>", lambda event = None: close())    
        top_wind.protocol("WM_DELETE_WINDOW", close)
    else:
        pass

window.bind("<Unmap>", lambda event = None: minimize())

window.mainloop()

